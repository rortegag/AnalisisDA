#########################################################################
#   Script para automatizar ánalisis de usuarios en Directorio Activo   #
#                                                                       #
#   Versión: 1.0                                                        #
#   Modificado por: ROG                                                 #
#########################################################################

from pathlib import Path
import glob
import csv
import os.path, time
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

def bitfield(n):
    return [1 if digit=='1' else 0 for digit in bin(n)[2:]]

def createSheets():
    wb.create_sheet("Deshabilitados-Bloqueados")
    wb.create_sheet("Contraseña no requerida")
    wb.create_sheet("Contraseña no expira")
    wb.create_sheet("Usuarios Inactivos")
    wb.create_sheet("Nunca han accedido")
    wb.create_sheet("No cambio contraseña")
    wb.create_sheet("Acceso y no cambio pass")
    wb.create_sheet("Usuarios Administradores")
    wb.create_sheet("Usuarios Genéricos")

    for sheet in wb:
        ws_active = sheet
        ws_active.append(listOfColumnNames)
        formatFirstRow(ws_active[1])

def formatFirstRow(firstRow):
    for cel in firstRow:
        cel.fill = PatternFill("solid", start_color="FF3F3F")
        cel.font = Font(bold=True)

def checkPassword(row, bitarray):
    if(bitarray[15] == 1):
        ws_active = wb["Contraseña no expira"]
        ws_active.append(row)
    else:
        if(row[listOfIndex[10]] != ""):
            pwdLastSet = datetime(1601, 1, 1) + timedelta(seconds=int(row[listOfIndex[10]])/10000000)
            lastLogon = getLastLogon()

            # Usuarios que están incumpliendo la política pero que puede que no hayan hecho login y por tanto no han podido cambiar contraseña
            if(datetime.fromtimestamp(creationTime)-pwdLastSet > timedelta(days=maxPasswordAge)):
                ws_active = wb["No cambio contraseña"]
                ws_active.append(row)

                # Si no está cumpliendo el cambio de contraseña y ha accedido está incumpliendo la política
                if(lastLogon != 0 and lastLogon > pwdLastSet):
                    ws_active = wb["Acceso y no cambio pass"]
                    ws_active.append(row)
        else:
            ws_active = wb["No cambio contraseña"]
            ws_active.append(row)
                
def getLastLogon():
    lastLogon = datetime(1601, 1, 1)
    lastLogonTimestamp = datetime(1601, 1, 1)
    if(row[listOfIndex[8]] == "" and row[listOfIndex[9]] == ""):
        return 0
    else:
        if(row[listOfIndex[8]] != "" and row[listOfIndex[9]] != ""):
            lastLogon = datetime(1601, 1, 1) + timedelta(seconds=int(row[listOfIndex[8]])/10000000)
            lastLogonTimestamp = datetime(1601, 1, 1) + timedelta(seconds=int(row[listOfIndex[9]])/10000000)
        elif(row[listOfIndex[8]] == "" and row[listOfIndex[9]] != ""):
            lastLogonTimestamp = datetime(1601, 1, 1) + timedelta(seconds=int(row[listOfIndex[9]])/10000000)
        elif(row[listOfIndex[8]] != "" and row[listOfIndex[9]] == ""):
            lastLogon = datetime(1601, 1, 1) + timedelta(seconds=int(row[listOfIndex[8]])/10000000)

        return max(lastLogon, lastLogonTimestamp)

def checkActivity():
    # Si lastLogon y lastLogonTimestamp están vacios nunca han accedido al sistema, en caso contrario comprobar último acceso
    lastLogonOk = getLastLogon()

    if(lastLogonOk == 0):
        ws_active = wb["Nunca han accedido"]
        ws_active.append(row)
    elif(datetime.fromtimestamp(creationTime)-lastLogonOk > timedelta(days=inactivityTime)):
        ws_active = wb["Usuarios Inactivos"]
        ws_active.append(row)

# Variables para definir inactividad (días desde el último login) y política de contraseñas en DIAS
inactivityTime = 150 # 150 días aprox. 5 mesesº
maxPasswordAge = 42 # 42 días es lo definido por defecto

currentPath = str(Path.cwd())

inputFiles = glob.glob(currentPath + "\input\*.csv")
print (inputFiles)

# Lista con los nombres de columnas que nos interesan
interestingColumnNames = ["DN", "objectClass", "cn", "distinguishedName", "whenCreated", "whenChanged", "name", "userAccountControl", "lastLogon", "lastLogonTimestamp", "pwdLastSet", "memberOf"]

# Admite varios csv, se procesarán todos
for file in inputFiles:
    creationTime = os.path.getctime(file) #Convertir a fecha con time.ctime
    print("File creation time: " + str(time.ctime(creationTime)))
    with open(str(file), newline='') as csvfile:
        # Leer el contenido completo del CSV y creamos un nuevo Excel sobre el que se escribiran los resultados
        csvContent = csv.reader(csvfile, delimiter=',')
        wb = Workbook()
        ws = wb.active
        ws.title = "Usuarios"

        # Guardamos la primera fila (Nombre de las columnas) en una lista y creamos una lista con los indices de las columnas que nos interesan. Se crean también las hojas del Excel
        listOfColumnNames = next(csvContent)
        createSheets()
        listOfIndex = []

        for x in interestingColumnNames:
            listOfIndex.append(listOfColumnNames.index(x))

        # Se recorre fila a fila el CSV        
        for row in csvContent:
            # Eliminar los objectClass = computer, se añaden todos los usuarios a la primera hoja del Excel resultado
            if(row[listOfIndex[1]] == "user"):
                ws.append(row)

                # Tratar el campo UserAccountControl
                bitarray = bitfield((int((row[listOfIndex[7]]))))
                while(len(bitarray) < 32):
                    bitarray.insert(0, 0)

                # Si el usuario está bloqueado (27) o deshabilitado (30) no nos interesa
                if(bitarray[30] == 1 or bitarray[27] == 21):
                    ws_active = wb["Deshabilitados-Bloqueados"]
                    ws_active.append(row)
                else:
                    # Si está activo pero no requiere contraseña no hacen falta más comprobaciones
                    if(bitarray[26] == 1):
                        ws_active = wb["Contraseña no requerida"]
                        ws_active.append(row)
                    else:
                        checkPassword(row, bitarray)
                    
                    # Usuarios inactivos
                    checkActivity()

                    # Si el campo memberOf contiene cualquiera de estas cadenas, será un usuario administrador de dominio
                    possibleValues = ["Admins. del dominio", "41646D696E732E2064656C20646F6D696E696F", "41646d696e", "446f6d61696e2041646d696e"] # Si se descubren nuevos valores se pueden añadir aquí
                    if(any(value in row[listOfIndex[11]] for value in possibleValues)):
                        ws_active = wb["Usuarios Administradores"]
                        ws_active.append(row)

                    # Usuarios genéricos
                    usersFile = open(currentPath + "\config\common-usernames")
                    if(row[listOfIndex[2]] in usersFile.read() or row[listOfIndex[6]] in usersFile.read()):
                        ws_active = wb["Usuarios Genéricos"]
                        ws_active.append(row)

        # Se guarda el Excel de resultados en la carpeta output
        wb.save(currentPath + "\output\\" + file.rsplit('\\', 1)[-1].rsplit('.', 1)[-2] + ".xlsx" )
        


